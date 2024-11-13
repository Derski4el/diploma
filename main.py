from openpyxl import load_workbook
from tabulate import tabulate

workbook = load_workbook("тестовые_оценки.xlsx")
sheet = workbook.active

table_data = []

for row in sheet.iter_rows(min_row=2, values_only=True):
    fio = str(row[1])
    grades = row[2:63]

    grade_counts = {
        "Пропуски": sum(1 for grade in grades if grade == 0),
        "Оценка 2": sum(1 for grade in grades if grade == 2),
        "Оценка 3": sum(1 for grade in grades if grade == 3),
        "Оценка 4": sum(1 for grade in grades if grade == 4),
        "Оценка 5": sum(1 for grade in grades if grade == 5)
    }
    table_data.append([fio] + list(grade_counts.values()))

headers = ["ФИО", "Пропуски", "Оценка 2", "Оценка 3", "Оценка 4", "Оценка 5"]
print(tabulate(table_data, headers=headers, tablefmt="pipe"))
