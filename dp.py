from openpyxl import load_workbook
from tabulate import tabulate
import os

folder_path = "C:/Users/4el/Desktop/diplom/ИСИП-101"
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

# Создаем словарь для хранения общих данных по каждому студенту
total_data = {}

for excel_file in excel_files:
    print(f"\nОбработка файла: {excel_file}")
    workbook = load_workbook(os.path.join(folder_path, excel_file))
    sheet = workbook.active
    
    for row in sheet.iter_rows(min_row=5, values_only=True):
        fio = str(row[0])
        grades = row[0:20]
        
        grade_counts = {
            "Пропуски": sum(1 for grade in grades if grade == "н"),
            "Оценка 2": sum(1 for grade in grades if grade == 2),
            "Оценка 3": sum(1 for grade in grades if grade == 3),
            "Оценка 4": sum(1 for grade in grades if grade == 4),
            "Оценка 5": sum(1 for grade in grades if grade == 5)
        }
        
        # Если студент уже есть в общих данных, складываем оценки
        if fio in total_data:
            for key in grade_counts:
                total_data[fio][key] += grade_counts[key]
        else:
            total_data[fio] = grade_counts

# Преобразуем общие данные в формат для вывода
final_table_data = []
for fio, grades in total_data.items():
    final_table_data.append([fio] + list(grades.values()))

headers = ["ФИО", "Пропуски", "Оценка 2", "Оценка 3", "Оценка 4", "Оценка 5"]
print("\nОбщая статистика по всем файлам:")
print(tabulate(final_table_data, headers=headers, tablefmt="pipe"))

def find_student(student_name, folder_path):
    print(f"\nПоиск информации по студенту: {student_name}")
    print("-" * 50)
    
    found = False
    for excel_file in excel_files:
        workbook = load_workbook(os.path.join(folder_path, excel_file))
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=5, values_only=True):
            if str(row[0]).lower() == student_name.lower():
                found = True
                print(f"\n{excel_file}:", end=" ")
                grades = row[0:20]
                for grade in grades:
                    print(grade, end=" ")
    
    if not found:
        print(f"Студент {student_name} не найден в журналах")

# После вывода общей таблицы добавляем:
while True:
    search = input("\nВведите ФИО студента для поиска (или 'выход' для завершения): ")
    if search.lower() == 'выход':
        break
    find_student(search, folder_path)